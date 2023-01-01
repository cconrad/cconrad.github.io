from openpyxl import load_workbook
# import requests as req

wb = load_workbook("/mnt/c/Users/mail/OneDrive/Personal/Internal/www.clausconrad.com.20221229.Sitemap.xlsx")
ws = wb.active

current_row = 2

while ws[f"A{current_row}"].value:
    if ws[f"F{current_row}"].value and ws[f"F{current_row}"].value.lower() == "x":
        url_from = ws[f"A{current_row}"].value
        url_from = url_from.replace("?", " ")
        url_to = ws[f"E{current_row}"].value
        status = ws[f"D{current_row}"].value
        print(f"{url_from} {url_to} {status}")
    current_row += 1

# wb.save("/mnt/c/Users/mail/OneDrive/Personal/Internal/www.clausconrad.com.20221229.Sitemap.xlsx")
